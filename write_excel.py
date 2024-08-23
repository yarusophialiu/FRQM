import pandas as pd
import re
import os
from utils import *
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt



def create_df(fps_categories, values_dict):
    base_columns = [360, 480, 720, 864, 1080]
    all_columns = base_columns * 10
    # row = ["bitrate"] + base_columns
    df = pd.DataFrame(columns=all_columns)
    df.loc[0] = [None] * len(all_columns)  # Initialize a new row with Nones

    # Insert the values into the DataFrame
    for fps in fps_categories:
        # print(f'fps{fps}')
        start_index = all_columns.index(360) + fps_categories.index(fps) * len(base_columns)
        end_index = start_index + len(base_columns)
        num_values = min(len(values_dict[fps]), len(base_columns))
        if DEBUG:
            print(f'start_index {start_index}')
            print(f'end_index {end_index}')
            print(f'values_dict[fps] {values_dict[fps]}')
        df.iloc[0, start_index:end_index] = values_dict[fps][:num_values] # row column  
    return df



def write_to_excel(excel_path, bitrate, fps_categories, fps_data, sheet_name):
    df = create_df(fps_categories, fps_data)
    df.insert(0, 'bitrate', bitrate)
    mode = 'w'  if not os.path.exists(excel_path) else 'a'

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='overlay') as writer:
        if sheet_name in writer.book.sheetnames:
            # Load existing Sheet X
            startrow = writer.book[sheet_name].max_row
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)



def create_sheet(excel_path, sheet_name):
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()
        print(f'Created a new workbook: {excel_path}')

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
        # print(f'Sheet "{sheet_name}" created.')                    
    wb.save(excel_path)



def get_fps_data(match, bitrate_data):
    bitrate = int(match[0])
    data = match[1].strip()
    fps_sections = re.split(r'========================= fps(\d+) =========================', data)
    if DEBUG:
        print(f'fps_sections \n {fps_sections}')
    fps_data = {}
    for i in range(1, len(fps_sections), 2):
        fps = int(fps_sections[i])
        cvvdp_values = re.findall(r'FRQM=([-]?\d+\.\d+)', fps_sections[i+1])
        if DEBUG:
            print(f'cvvdp_values \n {cvvdp_values}')              
        fps_data[fps] = list(map(float, cvvdp_values))


    bitrate_data[bitrate] = fps_data # fps_data is like values_dict

    fps_categories = list(fps_data.keys())
    # print(f'fps_categories {fps_categories}')
    # print(f'fps_data {fps_data}')
    fps_data = {key: value[1:] + [value[0]] for key, value in fps_data.items()}
    # print(f'bitrate {bitrate}, fps data \n {fps_data}\n\n\n')

    # print(f'fps_categories {list(fps_data.keys())}')
    fps_categories = sorted(fps_categories)
    # print(f'fps_categories {fps_categories}')
    return fps_categories, fps_data



def get_name(jobid):
    path, seg, speed = mapIdToPath(jobid-1)
    # print(f'path, seg, speed {path, seg, speed}\n')
    file_path = f'{cleaned_scene_path}/{SCENE}_{jobid}.txt'
    sheet_name = f'path{path}_seg{seg}_{speed}'  # Specify your desired sheet name here
    return file_path, sheet_name



def get_rows():
    base_columns = [360, 480, 720, 864, 1080]
    all_columns = base_columns * 10
    row1 = ["bitrate"] + all_columns

    fps_values = ["fps30", "fps40", "fps50", "fps60", "fps90", "fps100", "fps110", "fps120"]
    row2 = []

    fill_color = PatternFill(start_color="ffc0cb", end_color="ffc0cb", fill_type="solid")
    fps_positions = []
    for fps in fps_values:
        row2.append(fps)
        fps_positions.append(len(row2)+1)  # Track the position of each fps value
        row2.extend([" "] * 4)  # Add four empty spaces after each fps value
    row2 = [" "] + row2
    row2 = row2[:len(row1) + 1]
    return row1, row2, fps_positions, fill_color


def load_data_to_excel(excel_path, jobid_list):
    for jobid in jobid_list:
        print(f'====================== jobid {jobid} ======================')
        file_path, sheet_name = get_name(jobid)
        if WRITE_EXCEL:
            create_sheet(excel_path, sheet_name)
        with open(file_path, 'r') as file:
            content = file.read()

        pattern = r"========================= bitrate (\d+) =========================(.*?)((?========================== bitrate)|$)"
        matches = re.findall(pattern, content, re.DOTALL)
        bitrate_data = {}
        for match in matches:
            bitrate = int(match[0])
            # if DEBUG:
            print(f'============== {bitrate}kbps ==============')
            fps_categories, fps_data = get_fps_data(match, bitrate_data)
            score_values = list(fps_data.values())
            # print(f'bitrate {bitrate}, fps data \n {fps_data}\n\n\n')
            # print(f'bitrate {bitrate}, fps_categories \n {fps_categories}')
            # print(f'bitrate {bitrate}, score_values \n {score_values}\n\n\n')

            sorted_fps_data = dict(sorted(fps_data.items()))
            # Extracting sorted FPS values and corresponding scores
            fps_values = list(sorted_fps_data.keys())
            score_values = list(sorted_fps_data.values())
            print(f'sorted_fps_data \n {sorted_fps_data}')
            print(f'fps_values \n {fps_values}')
            print(f'score_values \n {score_values}')


            if PLOT:
                # Plotting
                plt.figure(figsize=(10, 6))
                plt.plot(fps_categories, score_values, marker='o', linestyle='-', color='b')

                path, seg, speed = mapIdToPath(jobid-1)
                plt.title(f'FRQM results- scene {SCENE} - reference fps 166 - path{path}_seg{seg}, speed {speed} - {bitrate} kbps')
                plt.xlabel('FPS')
                plt.ylabel('Score')
                plt.grid(True)
                plt.xticks(fps_categories) 
                # plt.show()
                
                p1 = f'plots-166'
                os.makedirs(p1, exist_ok=True)
                img_path = f"{p1}/p1_{SCENE}_{sheet_name}_{bitrate}.png"
                plt.savefig(img_path)
                # print(f"File saved: {img_path}")
                # if SAVE and (not os.path.exists(img_path)):
                #     plt.savefig(img_path)
                #     print(f"File saved: {img_path}")
                # else:
                #     print(f"File already exists: {img_path}")
                plt.show()

            if WRITE_EXCEL:
                write_to_excel(excel_path, bitrate, fps_categories, fps_data, sheet_name)


def append_header_rows(excel_path, jobid_list):
    wb = load_workbook(excel_path)
    row1, row2, fps_positions, fill_color = get_rows()

    for jobid in jobid_list:
        _, sheet_name = get_name(jobid)
        ws = wb[sheet_name]
        ws.append(row1)
        ws.append(row2)
        for pos in fps_positions:
            cell1 = ws.cell(row=ws.max_row-1, column=pos)
            cell2 = ws.cell(row=ws.max_row, column=pos)
            cell1.fill = fill_color
            cell2.fill = fill_color
        wb.save(excel_path)





# download cvvdp results from HPC
# process results by running this file
# write to excel using write_excel.py
# plot using plot_cvvdp.py

# close excel file to write data to it
if __name__ == "__main__":
    SCENES = ['bedroom', 'bistro', 'crytek_sponza', 'gallery', 'living_room', \
            'lost_empire', 'room', 'sibenik', 'suntemple', 'suntemple_statue']
    
    SCENES = ['bistro']
    
    # SCENE = "living_room"
    CLEANED_DIR = "08-22-166"
    PLOT = True
    SAVE = True
    WRITE_EXCEL = False
    DEBUG = False
    excel_dir = r'C:\Users\15142\Projects\VRR\VRR_Plot_FRQM'
    today = date.today()

    for SCENE in SCENES:
        print(f'=========================================== SCENE {SCENE} ===========================================')
        cleaned_scene_path = f'{CLEANED_DIR}/cleaned_{SCENE}'
        excel_path = '' # TODO: implement later
        # excel_path = f'{excel_dir}/data-{today}/{SCENE}.xlsx'
        # os.makedirs(f'{excel_dir}/data-{today}', exist_ok=True)
        jobid_list = [i for i in range(1, 2)] # TODO: change node
        # jobid_list = [i for i in range(18, 46)] # TODO: change node

        # TODO: loop load_data_to_excel
        load_data_to_excel(excel_path, jobid_list)
        # append_header_rows(excel_path, jobid_list)